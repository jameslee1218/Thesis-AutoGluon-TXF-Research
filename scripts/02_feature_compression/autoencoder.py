#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
技術指標 Autoencoder 壓縮訓練腳本
將同類技術指標壓縮為單一數值
"""

import sys
import io

# 設定輸出編碼為 UTF-8（解決 Windows 控制台編碼問題）
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# 專案 config（僅讀寫 data/）
from pathlib import Path as _Path
_REPO_ROOT = _Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))
import config as _config
DATA_DIR = str(_config.get_dataset_dir("0900"))
OUTPUT_DIR = str(_config.get_output_0900_dir())

import pandas as pd
import numpy as np
import os
import glob
from pathlib import Path
import warnings
from datetime import datetime
import json
import random

# 在導入 TensorFlow 之前設定環境變數（幫助 TensorFlow 找到 CUDA）
# 這可以幫助 TensorFlow 在 Windows 上找到 CUDA 庫
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '1'  # 減少 TensorFlow 日誌輸出
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'  # 關閉 oneDNN 選項

# 深度學習相關
import tensorflow as tf
# TensorFlow 2.10+ 中，keras 是獨立包，需要使用 keras.src
import keras
from keras.src import layers, models, callbacks
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_squared_error
from skopt import gp_minimize
from skopt.space import Integer, Real, Categorical
from skopt.utils import use_named_args

# 圖表和輸出
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # 使用非交互式後端
import openpyxl
from openpyxl.drawing.image import Image
import io

warnings.filterwarnings('ignore')

# ==================== 可調整參數 ====================
# 技術指標群組定義
INDICATOR_GROUPS = {
    "STOCH": ["STOCH_K_14", "STOCH_D_14"],
    "STOCHF": ["STOCHF_K_14", "STOCHF_D_14"],
    "STOCHRSI": ["STOCHRSI_K_14", "STOCHRSI_D_14"],
    "MACD": ["MACD_12_26", "MACD_signal_12_26", "MACD_hist_12_26"],
    "BBANDS": ["BBANDS_upper_20", "BBANDS_middle_20", "BBANDS_lower_20"],
    "ADX_DMI": ["ADX_14", "ADXR_14", "PDI_14", "MDI_14", "DX_14"],
    "AROON": ["AROON_Down_14", "AROON_Up_14", "AROONOSC_14"]
}

# 資料路徑（由上方 config 設定，僅 data/dataset/0900、data/output_0900）
# DATA_DIR, OUTPUT_DIR 已於檔頭設定

# 資料切分比例（維持）
VAL_SPLIT = 0.2  # 驗證集 20%
TEST_SPLIT = 0.1  # 測試集 10%

# Autoencoder 架構參數（維持；若想同步擴，建議再開新搜尋空間）
ENCODER_DIMS = [256, 128]  # 編碼器層次
DECODER_DIMS = [128, 256]  # 解碼器層次

# ==================== 超參數搜尋空間 ====================

# 擴大學習率候選（近似對數刻度）
LEARNING_RATES = [3e-4, 5e-4, 8e-4, 1e-3, 1.5e-3, 2e-3, 3e-3]

# 更細的 dropout 刻度（0~0.4 常見甜區）
DROPOUT_RATES = [0.0, 0.05, 0.1, 0.2, 0.3, 0.4]

# 擴大 batch size（依 GPU VRAM 視情況裁剪）
BATCH_SIZES = [128, 192, 256, 384, 512, 768, 1024]

# ==== 貝葉斯優化參數（加長探索） ====
BAYESIAN_N_CALLS = 32  # 原 12 → 48（可視資源 32~60 之間）

# 訓練參數（略微放寬上限、耐心值）
EARLY_STOPPING_PATIENCE = 16  # 原 12 → 16
MAX_EPOCHS = 300  # 原 200 → 300
RANDOM_SEED = 42

# 搜索階段的早停耐心（較短，加快搜索）
SEARCH_EARLY_STOPPING_PATIENCE = 8

# 優化選項
SKIP_FINAL_TRAINING = True  # 設為 True 可跳過最終訓練，直接使用搜索階段最佳模型（更快但可能性能稍差）

# 固定瓶頸層大小（全部壓成 1 維）
FIXED_BOTTLENECK = 1

# 改為連續取值（loguniform 分佈）而非固定清單
# Real 和 Integer 已在文件開頭導入，無需重複導入
LEARNING_RATE_SPACE = Real(3e-4, 2e-3, prior='log-uniform', name='learning_rate')

# dropout（0~0.4 常見甜區）
DROPOUT_RATES = [0.0, 0.05, 0.1, 0.2, 0.3, 0.4]

# batch size（依 GPU 記憶體調整）
BATCH_SIZES = [128, 192, 256, 384, 512, 768, 1024]

# ==== 貝葉斯優化參數 ====
BAYESIAN_N_CALLS = 32
EARLY_STOPPING_PATIENCE = 16
MAX_EPOCHS = 300
RANDOM_SEED = 42
SEARCH_EARLY_STOPPING_PATIENCE = 8
SKIP_FINAL_TRAINING = True


# ==================== 設定隨機種子 ====================
def set_random_seeds(seed):
    """設定所有隨機種子以確保可重現性"""
    random.seed(seed)
    np.random.seed(seed)
    tf.random.set_seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)

set_random_seeds(RANDOM_SEED)

# @title
# ==================== 設定隨機種子 ====================
def set_random_seeds(seed):
    """設定所有隨機種子以確保可重現性"""
    random.seed(seed)
    np.random.seed(seed)
    tf.random.set_seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)

set_random_seeds(RANDOM_SEED)

# ==================== 資料載入 ====================
def load_all_data(data_dir):
    """載入所有CSV檔案並合併，按年份分組"""
    print("=" * 60)
    print("[LOAD] 載入資料...")
    
    csv_files = glob.glob(os.path.join(data_dir, "TX*_1K_qlib_indicators_complete.csv"))
    csv_files.sort()  # 按檔名排序以確保時間順序
    
    print(f"找到 {len(csv_files)} 個CSV檔案")
    
    all_data = []
    for file in csv_files:
        try:
            df = pd.read_csv(file)
            df['datetime'] = pd.to_datetime(df['datetime'])
            all_data.append(df)
        except Exception as e:
            print(f"[WARN] 讀取檔案失敗: {os.path.basename(file)} - {e}")
    
    if not all_data:
        raise ValueError("[ERROR] 沒有成功載入任何資料！")
    
    combined_df = pd.concat(all_data, ignore_index=True)
    combined_df = combined_df.sort_values('datetime').reset_index(drop=True)
    
    # 提取年份
    combined_df['year'] = combined_df['datetime'].dt.year
    
    print(f"[OK] 成功載入資料，總共 {len(combined_df):,} 筆記錄")
    print(f"[INFO] 時間範圍: {combined_df['datetime'].min()} 至 {combined_df['datetime'].max()}")
    
    # 顯示年份分布
    year_counts = combined_df['year'].value_counts().sort_index()
    print(f"[INFO] 年份分布:")
    for year, count in year_counts.items():
        print(f"  {year}: {count:,} 筆")
    
    return combined_df

# ==================== 滾動窗口切分 ====================
def create_rolling_windows(df, train_years=2, compress_years=1):
    """創建滾動窗口：前N年訓練，後M年壓縮
    
    Args:
        df: 包含 'year' 和 'datetime' 列的 DataFrame
        train_years: 訓練年數（預設2年）
        compress_years: 壓縮年數（預設1年）
    
    Returns:
        windows: 列表，每個元素為 (train_df, compress_df, window_name)
    """
    years = sorted(df['year'].unique())
    windows = []
    
    print("=" * 60)
    print(f"[WINDOW] 創建滾動窗口（訓練: {train_years}年, 壓縮: {compress_years}年）...")
    print(f"可用年份: {years}")
    
    # 從第 train_years 年開始，每次滾動 compress_years 年
    i = train_years
    window_idx = 1
    
    while i + compress_years <= len(years):
        train_year_start = years[i - train_years]
        train_year_end = years[i - 1]
        compress_year_start = years[i]
        compress_year_end = years[i + compress_years - 1]
        
        # 訓練資料（前N年）
        train_df = df[(df['year'] >= train_year_start) & (df['year'] <= train_year_end)].copy()
        
        # 壓縮資料（後M年）
        compress_df = df[(df['year'] >= compress_year_start) & (df['year'] <= compress_year_end)].copy()
        
        if len(train_df) > 0 and len(compress_df) > 0:
            window_name = f"W{window_idx}_{train_year_start}-{train_year_end}_compress_{compress_year_start}-{compress_year_end}"
            windows.append((train_df, compress_df, window_name))
            
            print(f"  窗口 {window_idx}: 訓練 {train_year_start}-{train_year_end} ({len(train_df):,}筆), "
                  f"壓縮 {compress_year_start}-{compress_year_end} ({len(compress_df):,}筆)")
            window_idx += 1
        
        # 滾動到下一個窗口（每次前進 compress_years 年）
        i += compress_years
    
    print(f"[OK] 共創建 {len(windows)} 個滾動窗口")
    return windows

# 載入所有資料（不再在模組層級執行）
# df = load_all_data(DATA_DIR)  # 移到 main() 中執行

# @title
# ==================== 資料切分（用於訓練集內部切分） ====================
def time_split_data(df, val_split=0.2, test_split=0.1):
    """按時間順序切分資料（用於訓練集內部的 train/val/test 切分）"""
    n_total = len(df)
    n_test = int(n_total * test_split)
    n_val = int(n_total * val_split)
    n_train = n_total - n_val - n_test
    
    train_df = df.iloc[:n_train].copy()
    val_df = df.iloc[n_train:n_train+n_val].copy()
    test_df = df.iloc[n_train+n_val:].copy()
    
    print("=" * 60)
    print("[INFO] 資料切分結果 (Time-split):")
    print(f"  訓練集: {len(train_df):,} 筆 ({len(train_df)/n_total*100:.1f}%)")
    print(f"  驗證集: {len(val_df):,} 筆 ({len(val_df)/n_total*100:.1f}%)")
    print(f"  測試集: {len(test_df):,} 筆 ({len(test_df)/n_total*100:.1f}%)")
    print(f"  訓練集時間: {train_df['datetime'].min()} 至 {train_df['datetime'].max()}")
    print(f"  驗證集時間: {val_df['datetime'].min()} 至 {val_df['datetime'].max()}")
    print(f"  測試集時間: {test_df['datetime'].min()} 至 {test_df['datetime'].max()}")
    
    return train_df, val_df, test_df

# ==================== 資料準備 ====================
def prepare_indicator_data(df, indicator_cols):
    """準備指定指標的資料"""
    # 檢查欄位是否存在
    missing_cols = [col for col in indicator_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"[ERROR] 缺少欄位: {missing_cols}")
    
    data = df[indicator_cols].values
    # 處理無限大和NaN值
    data = np.nan_to_num(data, nan=0.0, posinf=1e6, neginf=-1e6)
    return data

# ==================== Autoencoder 模型 ====================
def build_autoencoder(input_dim, bottleneck_size, dropout_rate=0.0):
    """建立 Autoencoder 模型
    
    架構: Input → 256 → 128 → bottleneck → 128 → 256 → Output
    """
    input_layer = layers.Input(shape=(input_dim,), name='input')
    
    # 編碼器
    x = layers.Dense(ENCODER_DIMS[0], activation='relu', name='encoder_1')(input_layer)
    if dropout_rate > 0:
        x = layers.Dropout(dropout_rate)(x)
    x = layers.Dense(ENCODER_DIMS[1], activation='relu', name='encoder_2')(x)
    if dropout_rate > 0:
        x = layers.Dropout(dropout_rate)(x)
    bottleneck = layers.Dense(bottleneck_size, activation='relu', name='bottleneck')(x)
    
    # 解碼器
    x = layers.Dense(DECODER_DIMS[0], activation='relu', name='decoder_1')(bottleneck)
    if dropout_rate > 0:
        x = layers.Dropout(dropout_rate)(x)
    x = layers.Dense(DECODER_DIMS[1], activation='relu', name='decoder_2')(x)
    if dropout_rate > 0:
        x = layers.Dropout(dropout_rate)(x)
    output = layers.Dense(input_dim, activation='linear', name='output')(x)
    
    model = models.Model(inputs=input_layer, outputs=output, name='autoencoder')
    return model

# ==================== 訓練函數 ====================
def train_autoencoder(X_train, X_val, bottleneck_size, lr, dropout_rate, batch_size, 
                     max_epochs=200, patience=12, group_name="", show_progress=False):
    """訓練單一 Autoencoder"""
    input_dim = X_train.shape[1]
    
    # 建立模型
    model = build_autoencoder(input_dim, bottleneck_size, dropout_rate)
    
    # 編譯模型
    optimizer = keras.optimizers.Adam(learning_rate=lr)
    model.compile(optimizer=optimizer, loss='mse', metrics=['mse'])
    
    # 早停回調
    early_stopping = callbacks.EarlyStopping(
        monitor='val_loss',
        patience=patience,
        restore_best_weights=True,
        verbose=0
    )
    
    # 訓練歷史記錄
    history_callback = callbacks.History()
    
    # 自定義回調來顯示進度
    class ProgressCallback(callbacks.Callback):
        def on_epoch_end(self, epoch, logs=None):
            if show_progress and (epoch + 1) % 10 == 0:
                print(f"Epoch {epoch+1}/{max_epochs}, Loss: {logs['loss']:.6f}, Val Loss: {logs['val_loss']:.6f}", 
                      end='\r', flush=True)
    
    progress_callback = ProgressCallback() if show_progress else None
    
    # 記錄訓練開始時間
    train_start_time = datetime.now()
    
    # 訓練模型
    callbacks_list = [early_stopping, history_callback]
    if progress_callback:
        callbacks_list.append(progress_callback)
    
    history = model.fit(
        X_train, X_train,
        validation_data=(X_val, X_val),
        epochs=max_epochs,
        batch_size=batch_size,
        callbacks=callbacks_list,
        verbose=0
    )
    
    # 記錄訓練結束時間
    train_end_time = datetime.now()
    train_duration = (train_end_time - train_start_time).total_seconds()
    
    # 計算每個epoch的時間（平均）
    n_epochs = len(history.history['loss'])
    if n_epochs > 0:
        time_per_epoch = train_duration / n_epochs
        # 生成時間軸（累積時間）
        epoch_times = [time_per_epoch * (i + 1) for i in range(n_epochs)]
        history.history['epoch_times'] = epoch_times
        history.history['total_time'] = train_duration
    else:
        history.history['epoch_times'] = []
        history.history['total_time'] = 0
    
    return model, history.history

# ==================== 超參數搜尋（貝葉斯優化） ====================
def hyperparameter_search(X_train, X_val, group_name, indicator_cols):
    """使用貝葉斯優化進行超參數搜尋"""
    print(f"\n{'='*60}")
    print(f"[SEARCH] 開始貝葉斯優化超參數搜尋: {group_name}")
    print(f"   輸入維度: {len(indicator_cols)}")
    print(f"   優化迭代次數: {BAYESIAN_N_CALLS}")
    
    # 定義搜索空間
    # Bottleneck 固定為 1（將所有特徵壓縮成一維度）
    input_dim = len(indicator_cols)
    FIXED_BOTTLENECK = 1  # 固定值
    
    print(f"   Bottleneck 固定為: {FIXED_BOTTLENECK}（將 {input_dim} 維壓縮為 1 維）")
    
    # 確保所有維度都有至少兩個候選值（learning rate 使用連續空間，無需檢查）
    assert len(DROPOUT_RATES) >= 2, f"Dropout 候選必須至少 2 個，當前: {DROPOUT_RATES}"
    assert len(BATCH_SIZES) >= 2, f"Batch Size 候選必須至少 2 個，當前: {BATCH_SIZES}"
    
    # 搜索空間：只搜索 lr, dropout, batch（bottleneck 已固定）
    dimensions = [
        LEARNING_RATE_SPACE,  # learning rate（連續空間，名稱='learning_rate'）
        Integer(0, len(DROPOUT_RATES) - 1, name='dropout_idx'),  # dropout 索引
        Integer(0, len(BATCH_SIZES) - 1, name='batch_idx'),  # batch size 索引
    ]
    
    # 儲存所有評估結果
    results = []
    best_val_mse = float('inf')
    best_config = None
    best_model = None
    best_history = None
    iteration_count = [0]  # 使用列表以便在嵌套函數中修改
    
    search_start_time = datetime.now()
    
    # 定義目標函數（貝葉斯優化要最小化的函數）
    @use_named_args(dimensions=dimensions)
    def objective(learning_rate, dropout_idx, batch_idx):
        """目標函數：返回驗證集 MSE（要最小化）
        
        參數：
            learning_rate: 學習率（連續值，來自 LEARNING_RATE_SPACE）
            dropout_idx: dropout 索引（離散值）
            batch_idx: batch size 索引（離散值）
        """
        iteration_count[0] += 1
        idx = iteration_count[0]
        
        # 將索引轉換為實際值（bottleneck 固定為 1）
        bottleneck = FIXED_BOTTLENECK
        lr = float(learning_rate)  # 學習率是連續值，直接使用
        dropout = DROPOUT_RATES[int(dropout_idx)]
        batch = BATCH_SIZES[int(batch_idx)]
        
        config = {
            'bottleneck': bottleneck,
            'lr': lr,
            'dropout': dropout,
            'batch': batch
        }
        
        # 計算進度
        progress = (idx - 1) / BAYESIAN_N_CALLS * 100
        elapsed_time = (datetime.now() - search_start_time).total_seconds()
        
        print(f"\n  [貝葉斯優化 {progress:.1f}%] [{idx}/{BAYESIAN_N_CALLS}] 測試組合:")
        print(f"    Bottleneck: {bottleneck}, LR: {lr:.0e}, "
              f"Dropout: {dropout}, Batch: {batch}")
        
        if idx > 1:
            avg_time = elapsed_time / (idx - 1)
            remaining = avg_time * (BAYESIAN_N_CALLS - idx + 1)
            print(f"    已用時間: {elapsed_time:.1f}秒 | 預計剩餘: {remaining:.1f}秒")
        
        try:
            # 訓練模型（搜索階段使用較短的耐心值）
            print(f"    [訓練中...] ", end='', flush=True)
            model, history = train_autoencoder(
                X_train, X_val,
                bottleneck_size=bottleneck,
                lr=lr,
                dropout_rate=dropout,
                batch_size=batch,
                max_epochs=MAX_EPOCHS,
                patience=SEARCH_EARLY_STOPPING_PATIENCE,  # 搜索階段使用較短的耐心值（8）
                group_name=group_name
            )
            
            # 使用訓練歷史中的最佳驗證損失作為目標值（更準確且更快）
            best_val = float(np.min(history['val_loss']))
            
            # 顯示訓練時間和結果
            if 'total_time' in history:
                epochs = len(history.get('loss', []))
                print(f"[完成] 訓練時間: {history['total_time']:.2f}秒 ({epochs} epochs)")
            
            print(f"    [結果] 最佳 Val Loss: {best_val:.6f}")
            
            # 更新最佳模型（bottleneck 固定為 1，只需比較驗證損失）
            nonlocal best_val_mse, best_config, best_model, best_history
            is_better = False
            if best_val < best_val_mse * 0.99:  # 明顯更好（>1%）
                is_better = True
            elif best_config is not None and best_val <= best_val_mse * 1.01:
                # 接近（±1%）且相同或更好（bottleneck 固定，無需比較）
                is_better = True
            
            # 只有當 is_better 成立時，才計算詳細的 MSE（用於報表，節省時間）
            train_mse = None
            val_mse = None
            if is_better:
                print(f"[評估中...] ", end='', flush=True)
                val_pred = model.predict(X_val, verbose=0)
                val_mse = mean_squared_error(X_val, val_pred)
                
                train_pred = model.predict(X_train, verbose=0)
                train_mse = mean_squared_error(X_train, train_pred)
                
                print(f"[完成] Train MSE: {train_mse:.6f}, Val MSE: {val_mse:.6f}")
                
                best_val_mse = best_val  # 使用 val_loss 作為比較基準
                best_config = config
                best_model = model
                best_history = history
                print(f"    [BEST] 更新最佳模型！")
            
            # 儲存結果（只保存基本信息，避免保存所有模型）
            result = {
                'config': config,
                'train_mse': train_mse,  # 可能為 None（只有 is_better 時才計算）
                'val_mse': val_mse,  # 可能為 None（只有 is_better 時才計算）
                'best_val_loss': best_val,  # 使用這個作為主要指標（從 history 中取得）
                'model': model if is_better else None,  # 只保存最佳模型以節省記憶體
                'history': history
            }
            results.append(result)
            
            # 返回最佳驗證損失（貝葉斯優化要最小化的值）
            return best_val
        
        except Exception as e:
            print(f"    [ERROR] 訓練失敗: {e}")
            # 返回一個很大的值，表示這組參數不好
            return 1e10
    
    # 執行貝葉斯優化
    print(f"\n[INFO] 開始貝葉斯優化（高斯過程）...")
    result_optimization = gp_minimize(
        func=objective,
        dimensions=dimensions,
        n_calls=BAYESIAN_N_CALLS,
        random_state=RANDOM_SEED,
        n_initial_points=min(4, BAYESIAN_N_CALLS),  # 初始隨機採樣點數
        acq_func='EI',  # Expected Improvement 採 acquisition function
        verbose=False
    )
    
    # 檢查是否有成功的結果
    if best_config is None:
        raise ValueError(f"[ERROR] {group_name} 沒有成功的訓練結果！")
    
    print(f"\n{'='*60}")
    print(f"[BEST] 最佳超參數 ({group_name}):")
    print(f"   Bottleneck: {FIXED_BOTTLENECK} (固定)")
    print(f"   Learning Rate: {best_config['lr']:.0e}")
    print(f"   Dropout: {best_config['dropout']}")
    print(f"   Batch Size: {best_config['batch']}")
    print(f"   最佳驗證損失: {best_val_mse:.6f}")
    # 如果有計算過的 val_mse，顯示它
    best_result = next((r for r in results if r.get('val_mse') is not None and r['config'] == best_config), None)
    if best_result:
        print(f"   驗證集 MSE: {best_result['val_mse']:.6f}")
    print(f"\n[INFO] 貝葉斯優化找到的最佳目標值: {result_optimization.fun:.6f}")
    print(f"[INFO] 最佳參數位置: {result_optimization.x}")
    
    return best_model, best_config, best_history, results

# ==================== 最終訓練 ====================
def final_training(X_train, X_val, X_test, best_config, group_name, best_model=None):
    """使用最佳參數在 Train+Val 上重訓，評估 Test
    
    如果提供 best_model，使用遷移學習（繼續訓練）而非完全重新訓練，節省時間
    """
    print(f"\n{'='*60}")
    print(f"[FINAL] 最終訓練: {group_name}")
    print(f"   使用 Train+Val 資料重新訓練...")
    
    # 合併 Train 和 Val
    X_train_val = np.vstack([X_train, X_val])
    
    # 切分 Train+Val 為新的 train 和 val（用於早停，比例為 80:20）
    n_val_final = int(len(X_train_val) * 0.2)
    X_train_final = X_train_val[:-n_val_final]
    X_val_final = X_train_val[-n_val_final:]
    
    print(f"   最終訓練集: {len(X_train_final):,} 筆")
    print(f"   最終驗證集: {len(X_val_final):,} 筆（用於早停）")
    print(f"   測試集: {len(X_test):,} 筆")
    
    # 如果提供了最佳模型，使用遷移學習（繼續訓練）而非完全重新訓練
    if best_model is not None:
        print(f"   [優化] 使用遷移學習：從搜索階段最佳模型繼續訓練（節省時間）")
        input_dim = X_train_final.shape[1]
        
        # 建立相同架構的新模型
        model = build_autoencoder(
            input_dim, 
            best_config['bottleneck'], 
            best_config['dropout']
        )
        
        # 編譯模型
        optimizer = keras.optimizers.Adam(learning_rate=best_config['lr'])
        model.compile(optimizer=optimizer, loss='mse', metrics=['mse'])
        
        # 複製最佳模型的權重（遷移學習）
        try:
            # 嘗試複製權重層對應
            best_layers = best_model.layers
            new_layers = model.layers
            
            # 複製可訓練層的權重
            for best_layer, new_layer in zip(best_layers, new_layers):
                if len(best_layer.get_weights()) > 0 and len(new_layer.get_weights()) > 0:
                    # 檢查層結構是否匹配
                    if (best_layer.get_weights()[0].shape == new_layer.get_weights()[0].shape and
                        len(best_layer.get_weights()) == len(new_layer.get_weights())):
                        new_layer.set_weights(best_layer.get_weights())
            print(f"   ✅ 權重遷移成功")
        except Exception as e:
            print(f"   ⚠️ 權重遷移失敗，將從頭訓練: {e}")
        
        # 早停回調
        early_stopping = callbacks.EarlyStopping(
            monitor='val_loss',
            patience=EARLY_STOPPING_PATIENCE,
            restore_best_weights=True,
            verbose=0
        )
        
        history_callback = callbacks.History()
        
        # 記錄訓練開始時間
        train_start_time = datetime.now()
        
        # 繼續訓練（通常只需要很少的 epochs，因為已經有好的初始權重）
        history = model.fit(
            X_train_final, X_train_final,
            validation_data=(X_val_final, X_val_final),
            epochs=MAX_EPOCHS,
            batch_size=best_config['batch'],
            callbacks=[early_stopping, history_callback],
            verbose=0,
            initial_epoch=0
        )
        
        # 記錄訓練結束時間
        train_end_time = datetime.now()
        train_duration = (train_end_time - train_start_time).total_seconds()
        
        # 計算每個epoch的時間
        n_epochs = len(history.history['loss'])
        if n_epochs > 0:
            time_per_epoch = train_duration / n_epochs
            epoch_times = [time_per_epoch * (i + 1) for i in range(n_epochs)]
            history.history['epoch_times'] = epoch_times
            history.history['total_time'] = train_duration
        else:
            history.history['epoch_times'] = []
            history.history['total_time'] = 0
        
        history = history.history
    else:
        # 完全重新訓練（原始方法）
        model, history = train_autoencoder(
            X_train_final, X_val_final,
            bottleneck_size=best_config['bottleneck'],
            lr=best_config['lr'],
            dropout_rate=best_config['dropout'],
            batch_size=best_config['batch'],
            max_epochs=MAX_EPOCHS,
            patience=EARLY_STOPPING_PATIENCE,
            group_name=group_name
        )
    
    epochs = len(history.get('loss', []))
    train_time = history.get('total_time', 0)
    print(f"[完成] 訓練時間: {train_time:.2f}秒 ({epochs} epochs)")
    
    print(f"   [評估中...] ", end='', flush=True)
    
    # 評估所有集合（使用完整的 Train+Val 和 Test）
    train_val_pred = model.predict(X_train_val, verbose=0)
    test_pred = model.predict(X_test, verbose=0)
    
    train_val_mse = mean_squared_error(X_train_val, train_val_pred)
    test_mse = mean_squared_error(X_test, test_pred)
    
    # 分別計算 Train 和 Val 的 MSE（僅用於報告）
    train_pred_only = model.predict(X_train, verbose=0)
    val_pred_only = model.predict(X_val, verbose=0)
    train_mse = mean_squared_error(X_train, train_pred_only)
    val_mse = mean_squared_error(X_val, val_pred_only)
    
    print(f"[完成]")
    print(f"   [結果] Train MSE: {train_mse:.6f}")
    print(f"   [結果] Val MSE: {val_mse:.6f}")
    print(f"   [結果] Train+Val MSE: {train_val_mse:.6f}")
    print(f"   [結果] Test MSE: {test_mse:.6f}")
    
    return model, {
        'train_mse': train_mse,
        'val_mse': val_mse,
        'train_val_mse': train_val_mse,
        'test_mse': test_mse,
        'history': history
    }

# ==================== 壓縮並保存資料 ====================
def compress_and_save_data(model, scaler, df, indicator_cols, group_name, output_dir, bottleneck_size, window_name=""):
    """使用訓練好的模型壓縮資料並保存為時間序列格式
    
    Args:
        model: 訓練好的 autoencoder 模型
        scaler: 標準化器
        df: 要壓縮的資料 DataFrame
        indicator_cols: 指標欄位列表
        group_name: 指標群組名稱
        output_dir: 輸出目錄
        bottleneck_size: 瓶頸層大小
        window_name: 窗口名稱（用於區分不同窗口的結果）
    """
    # 創建壓縮資料輸出目錄
    compressed_dir = os.path.join(output_dir, "compressed_data")
    os.makedirs(compressed_dir, exist_ok=True)
    
    # 準備所有資料
    print(f"   準備壓縮資料...")
    all_data = prepare_indicator_data(df, indicator_cols)
    
    # 標準化
    all_data_scaled = scaler.transform(all_data)
    
    # 提取編碼器部分（從輸入到 bottleneck）
    # 構建編碼器模型
    encoder_input = model.input
    encoder_output = None
    
    # 找到 bottleneck 層的輸出
    for layer in model.layers:
        if layer.name == 'bottleneck':
            encoder_output = layer.output
            break
    
    if encoder_output is None:
        # 如果找不到，使用最後一個編碼器層
        # 找到 bottleneck 之前的層
        bottleneck_layer_idx = None
        for i, layer in enumerate(model.layers):
            if layer.name == 'bottleneck':
                bottleneck_layer_idx = i
                break
        
        if bottleneck_layer_idx is not None:
            encoder_output = model.layers[bottleneck_layer_idx].output
        else:
            raise ValueError("無法找到編碼器輸出層")
    
    # 創建編碼器模型
    encoder_model = models.Model(inputs=encoder_input, outputs=encoder_output)
    
    # 使用編碼器壓縮資料
    print(f"   使用編碼器壓縮資料...")
    compressed_data = encoder_model.predict(all_data_scaled, verbose=0)
    
    # 創建包含 datetime 和壓縮特徵的 DataFrame
    compressed_df = pd.DataFrame(
        compressed_data,
        columns=[f"{group_name}_compressed_{i}" for i in range(bottleneck_size)]
    )
    
    # 添加 datetime 列（如果原始資料有）
    if 'datetime' in df.columns:
        compressed_df['datetime'] = df['datetime'].values
        # 將 datetime 移到第一列
        cols = ['datetime'] + [col for col in compressed_df.columns if col != 'datetime']
        compressed_df = compressed_df[cols]
    
    # 保存為 CSV（如果提供窗口名稱，加入文件名）
    if window_name:
        output_path = os.path.join(compressed_dir, f"{group_name}_{window_name}_compressed.csv")
    else:
        output_path = os.path.join(compressed_dir, f"{group_name}_compressed.csv")
    compressed_df.to_csv(output_path, index=False, encoding='utf-8-sig')
    
    # 顯示壓縮統計
    original_size = all_data.shape[1]
    compressed_size = bottleneck_size
    compression_ratio = original_size / compressed_size
    
    print(f"   ✅ 壓縮完成！")
    print(f"   原始維度: {original_size}")
    print(f"   壓縮後維度: {compressed_size}")
    print(f"   壓縮比: {compression_ratio:.2f}:1")
    print(f"   資料筆數: {len(compressed_df):,}")
    print(f"   保存路徑: {output_path}")
    
    return output_path

# ==================== 繪圖函數 ====================
def plot_training_history(history, group_name, output_path):
    """繪製訓練歷史（包含時間軸）"""
    # 如果有時間信息，創建3個子圖，否則2個
    has_time = 'epoch_times' in history and len(history['epoch_times']) > 0
    
    if has_time:
        fig, axes = plt.subplots(1, 3, figsize=(20, 5))
        epoch_times = history['epoch_times']
    else:
        fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    
    # MSE 圖（按 Epoch）
    axes[0].plot(history['loss'], label='Train MSE', linewidth=2)
    axes[0].plot(history['val_loss'], label='Val MSE', linewidth=2)
    axes[0].set_xlabel('Epoch', fontsize=12)
    axes[0].set_ylabel('MSE', fontsize=12)
    axes[0].set_title(f'{group_name} - Training MSE (by Epoch)', fontsize=14, fontweight='bold')
    axes[0].legend(fontsize=10)
    axes[0].grid(True, alpha=0.3)
    
    # 對數尺度 MSE 圖（按 Epoch）
    axes[1].semilogy(history['loss'], label='Train MSE', linewidth=2)
    axes[1].semilogy(history['val_loss'], label='Val MSE', linewidth=2)
    axes[1].set_xlabel('Epoch', fontsize=12)
    axes[1].set_ylabel('MSE (log scale)', fontsize=12)
    axes[1].set_title(f'{group_name} - Training MSE (Log Scale, by Epoch)', fontsize=14, fontweight='bold')
    axes[1].legend(fontsize=10)
    axes[1].grid(True, alpha=0.3)
    
    # 如果有時間信息，添加時間軸圖
    if has_time:
        # MSE 圖（按時間）
        axes[2].plot(epoch_times, history['loss'], label='Train MSE', linewidth=2)
        axes[2].plot(epoch_times, history['val_loss'], label='Val MSE', linewidth=2)
        axes[2].set_xlabel('Training Time (seconds)', fontsize=12)
        axes[2].set_ylabel('MSE', fontsize=12)
        axes[2].set_title(f'{group_name} - Training MSE (by Time)\nTotal: {history.get("total_time", 0):.1f}s', 
                         fontsize=14, fontweight='bold')
        axes[2].legend(fontsize=10)
        axes[2].grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    return output_path

# ==================== Excel 輸出 ====================
def save_results_to_excel(all_results, output_dir):
    """將所有結果保存到 Excel"""
    print(f"\n{'='*60}")
    print("[SAVE] 保存結果到 Excel...")
    
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, f"autoencoder_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    
    wb = openpyxl.Workbook()
    
    # 1. 摘要表
    ws_summary = wb.active
    ws_summary.title = "摘要"
    ws_summary.append(["技術指標群組", "輸入維度", "最佳 Bottleneck", "最佳 LR", "最佳 Dropout", 
                      "最佳 Batch Size", "Train MSE", "Val MSE", "Train+Val MSE", "Test MSE",
                      "訓練時間(秒)", "總訓練時間(秒)"])
    
    for group_name, result in all_results.items():
        config = result['best_config']
        final_scores = result['final_scores']
        history = final_scores.get('history', {})
        search_time = history.get('total_time', 0) if 'total_time' in history else 0
        
        # 計算總訓練時間（搜尋階段 + 最終訓練）
        final_history = final_scores.get('history', {})
        final_time = final_history.get('total_time', 0) if 'total_time' in final_history else 0
        
        # 計算搜尋階段的總時間
        search_total_time = 0
        for search_result in result.get('search_results', []):
            if isinstance(search_result, dict) and 'history' in search_result:
                search_total_time += search_result['history'].get('total_time', 0)
        
        total_training_time = search_total_time + final_time
        
        ws_summary.append([
            group_name,
            result['input_dim'],
            config['bottleneck'],
            config['lr'],
            config['dropout'],
            config['batch'],
            f"{final_scores['train_mse']:.6f}",
            f"{final_scores['val_mse']:.6f}",
            f"{final_scores.get('train_val_mse', final_scores['train_mse']):.6f}",
            f"{final_scores['test_mse']:.6f}",
            f"{final_time:.2f}",
            f"{total_training_time:.2f}"
        ])
    
    # 2. 為每個群組創建詳細工作表
    for group_name, result in all_results.items():
        ws = wb.create_sheet(title=group_name[:31])  # Excel工作表名稱限制31字元
        
        # 超參數搜尋結果
        ws.append(["超參數搜尋結果"])
        ws.append(["Bottleneck", "Learning Rate", "Dropout", "Batch Size", "Train MSE", "Val MSE"])
        
        for search_result in result['search_results']:
            config = search_result['config']
            train_mse = search_result.get('train_mse')
            val_mse = search_result.get('val_mse')
            ws.append([
                config['bottleneck'],
                config['lr'],
                config['dropout'],
                config['batch'],
                f"{train_mse:.6f}" if train_mse is not None else "N/A",
                f"{val_mse:.6f}" if val_mse is not None else "N/A"
            ])
        
        ws.append([])
        ws.append(["最佳配置"])
        best_config = result['best_config']
        ws.append(["Bottleneck", best_config['bottleneck']])
        ws.append(["Learning Rate", best_config['lr']])
        ws.append(["Dropout", best_config['dropout']])
        ws.append(["Batch Size", best_config['batch']])
        
        ws.append([])
        ws.append(["最終評估結果"])
        final_scores = result['final_scores']
        ws.append(["Train MSE", f"{final_scores['train_mse']:.6f}"])
        ws.append(["Val MSE", f"{final_scores['val_mse']:.6f}"])
        if 'train_val_mse' in final_scores:
            ws.append(["Train+Val MSE", f"{final_scores['train_val_mse']:.6f}"])
        ws.append(["Test MSE", f"{final_scores['test_mse']:.6f}"])
        
        # 添加時間信息
        ws.append([])
        ws.append(["訓練時間信息"])
        final_history = final_scores.get('history', {})
        if 'total_time' in final_history:
            ws.append(["最終訓練時間", f"{final_history['total_time']:.2f} 秒"])
            ws.append(["平均每 Epoch 時間", f"{final_history['total_time'] / max(len(final_history.get('loss', [])), 1):.2f} 秒"])
        else:
            ws.append(["最終訓練時間", "未記錄"])
        
        # 計算搜尋階段的總時間
        search_total_time = 0
        for search_result in result.get('search_results', []):
            if 'history' in search_result and 'total_time' in search_result['history']:
                search_total_time += search_result['history']['total_time']
        if search_total_time > 0:
            ws.append(["超參數搜尋總時間", f"{search_total_time:.2f} 秒"])
            ws.append(["總訓練時間", f"{search_total_time + final_history.get('total_time', 0):.2f} 秒"])
        
        # 插入圖片
        img_path = result['plot_path']
        if os.path.exists(img_path):
            try:
                img = Image(img_path)
                img.width = 800
                img.height = 300
                ws.add_image(img, f'A{ws.max_row + 3}')
            except Exception as e:
                print(f"  [WARN] 無法插入圖片 {img_path}: {e}")
    
    # 3. 訓練日誌
    ws_log = wb.create_sheet(title="訓練日誌")
    ws_log.append(["時間", "群組", "階段", "訊息"])
    
    for group_name, result in all_results.items():
        if 'log' in result:
            for log_entry in result['log']:
                ws_log.append(log_entry)
    
    wb.save(excel_path)
    print(f"[OK] Excel 已保存: {excel_path}")
    
    return excel_path

# ==================== 主程式 ====================
def main():
    """主程式"""
    print("=" * 60)
    print("[START] 技術指標 Autoencoder 壓縮訓練")
    print("=" * 60)
    print(f"[TIME] 開始時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"[SEED] 隨機種子: {RANDOM_SEED}")
    
    # GPU 檢查和配置
    print("\n[GPU] GPU 檢查和配置...")
    print(f"TensorFlow 版本: {tf.__version__}")
    print(f"TensorFlow 是否構建時包含 CUDA 支持: {tf.test.is_built_with_cuda()}")
    
    gpus = tf.config.list_physical_devices('GPU')
    print(f"可用 GPU 清單: {gpus}")
    
    if len(gpus) > 0:
        print(f"✅ 檢測到 {len(gpus)} 個 GPU 設備")
        try:
            # 啟用 GPU 記憶體增長（避免一次性分配所有記憶體）
            for gpu in gpus:
                tf.config.experimental.set_memory_growth(gpu, True)
            print("✅ GPU 記憶體增長已啟用")
            
            # 驗證 GPU 是否可用
            logical_gpus = tf.config.list_logical_devices('GPU')
            if len(logical_gpus) > 0:
                print(f"✅ GPU 可用於 TensorFlow 運算: {logical_gpus}")
                
                # 測試 GPU 運算
                try:
                    with tf.device('/GPU:0'):
                        test_tensor = tf.constant([1.0, 2.0, 3.0])
                        result = tf.reduce_sum(test_tensor)
                    print(f"✅ GPU 運算測試成功: {result.numpy()}")
                    print("TensorFlow 是否使用 GPU: True")
                except Exception as e:
                    print(f"⚠️ GPU 運算測試失敗: {e}")
                    print("TensorFlow 是否使用 GPU: False")
            else:
                print("❌ GPU 不可用於 TensorFlow 運算")
                print("TensorFlow 是否使用 GPU: False")
        except RuntimeError as e:
            print(f"⚠️ GPU 設定警告: {e}")
            print("TensorFlow 是否使用 GPU: False")
    else:
        print("❌ 沒有檢測到 GPU 設備")
        if not tf.test.is_built_with_cuda():
            print("⚠️ TensorFlow 當前版本似乎不包含 CUDA 支持（CPU-only 構建）")
            print("💡 提示：如果已安裝 CUDA，可能需要：")
            print("   1. 確保已安裝完整的 CUDA Toolkit（不僅是驅動）")
            print("   2. 安裝對應版本的 cuDNN")
            print("   3. 或考慮使用 conda 安裝支持 GPU 的 TensorFlow")
        print("將使用 CPU 進行運算")
        print("TensorFlow 是否使用 GPU: False")
    
    print()
    
    # 載入所有資料
    df = load_all_data(DATA_DIR)
    
    # 創建滾動窗口（前2年訓練，後1年壓縮）
    windows = create_rolling_windows(df, train_years=2, compress_years=1)
    
    if len(windows) == 0:
        raise ValueError("[ERROR] 無法創建任何滾動窗口！請檢查資料年份是否足夠。")
    
    # 創建輸出目錄
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # 儲存所有結果（按窗口和群組）
    all_results = {}
    
    # 對每個滾動窗口進行處理
    total_windows = len(windows)
    overall_start_time = datetime.now()
    
    for window_idx, (train_df_window, compress_df_window, window_name) in enumerate(windows, 1):
        print(f"\n{'='*80}")
        print(f"[WINDOW {window_idx}/{total_windows}] 處理窗口: {window_name}")
        print(f"{'='*80}")
        
        # 計算已用時間和預計剩餘時間
        elapsed = (datetime.now() - overall_start_time).total_seconds()
        if window_idx > 1:
            avg_time_per_window = elapsed / (window_idx - 1)
            remaining_windows = total_windows - window_idx + 1
            remaining_time = avg_time_per_window * remaining_windows
            print(f"已用時間: {elapsed:.1f}秒 | 預計剩餘: {remaining_time:.1f}秒 ({remaining_time/60:.1f}分鐘)")
        
        # 為當前窗口創建輸出目錄
        window_output_dir = os.path.join(OUTPUT_DIR, window_name)
        os.makedirs(window_output_dir, exist_ok=True)
        models_dir = os.path.join(window_output_dir, "models")
        plots_dir = os.path.join(window_output_dir, "plots")
        os.makedirs(models_dir, exist_ok=True)
        os.makedirs(plots_dir, exist_ok=True)
        
        # 對訓練資料進行內部切分（train/val/test）
        train_df, val_df, test_df = time_split_data(train_df_window, VAL_SPLIT, TEST_SPLIT)
        
        # 對每個技術指標群組進行處理
        total_groups = len(INDICATOR_GROUPS)
        window_start_time = datetime.now()
        
        for group_idx, (group_name, indicator_cols) in enumerate(INDICATOR_GROUPS.items(), 1):
            print(f"\n{'#'*60}")
            print(f"[WINDOW {window_idx}/{total_windows}] [GROUP {group_idx}/{total_groups}] 處理技術指標群組: {group_name}")
            print(f"   包含指標: {', '.join(indicator_cols)}")
            print(f"   窗口進度: {group_idx}/{total_groups} ({group_idx/total_groups*100:.1f}%)")
            
            # 計算已用時間和預計剩餘時間
            elapsed_group = (datetime.now() - window_start_time).total_seconds()
            if group_idx > 1:
                avg_time_per_group = elapsed_group / (group_idx - 1)
                remaining_groups = total_groups - group_idx + 1
                remaining_time = avg_time_per_group * remaining_groups
                print(f"   窗口內已用時間: {elapsed_group:.1f}秒 | 預計剩餘: {remaining_time:.1f}秒 ({remaining_time/60:.1f}分鐘)")
            
            print(f"{'#'*60}")
        
            try:
                # 準備資料
                X_train_raw = prepare_indicator_data(train_df, indicator_cols)
                X_val_raw = prepare_indicator_data(val_df, indicator_cols)
                X_test_raw = prepare_indicator_data(test_df, indicator_cols)
                
                # 標準化（只在訓練集上 fit）
                scaler = StandardScaler()
                X_train = scaler.fit_transform(X_train_raw)
                X_val = scaler.transform(X_val_raw)
                X_test = scaler.transform(X_test_raw)
                
                print(f"[OK] 資料準備完成")
                print(f"   訓練集形狀: {X_train.shape}")
                print(f"   驗證集形狀: {X_val.shape}")
                print(f"   測試集形狀: {X_test.shape}")
                
                # 超參數搜尋
                best_model, best_config, best_history, search_results = hyperparameter_search(
                    X_train, X_val, group_name, indicator_cols
                )
                
                # 保存搜尋階段的模型（可選）
                search_model_path = os.path.join(models_dir, f"{group_name}_search_best.h5")
                best_model.save(search_model_path)
            
                # 最終訓練（可選，如果 SKIP_FINAL_TRAINING=True 則跳過以節省時間）
                if SKIP_FINAL_TRAINING:
                    print(f"\n[OPTIMIZE] 跳過最終訓練，直接使用搜索階段最佳模型（節省時間）")
                    final_model = best_model
                    
                    # 使用搜索階段的歷史記錄
                    final_history = best_history
                    
                    # 評估最終模型（在完整數據集上）
                    print(f"   [評估中...] ", end='', flush=True)
                    X_train_val = np.vstack([X_train, X_val])
                    
                    train_val_pred = final_model.predict(X_train_val, verbose=0)
                    test_pred = final_model.predict(X_test, verbose=0)
                    
                    train_val_mse = mean_squared_error(X_train_val, train_val_pred)
                    test_mse = mean_squared_error(X_test, test_pred)
                    
                    train_pred_only = final_model.predict(X_train, verbose=0)
                    val_pred_only = final_model.predict(X_val, verbose=0)
                    train_mse = mean_squared_error(X_train, train_pred_only)
                    val_mse = mean_squared_error(X_val, val_pred_only)
                    
                    print(f"[完成]")
                    print(f"   [結果] Train MSE: {train_mse:.6f}")
                    print(f"   [結果] Val MSE: {val_mse:.6f}")
                    print(f"   [結果] Train+Val MSE: {train_val_mse:.6f}")
                    print(f"   [結果] Test MSE: {test_mse:.6f}")
                    
                    final_scores = {
                        'train_mse': train_mse,
                        'val_mse': val_mse,
                        'train_val_mse': train_val_mse,
                        'test_mse': test_mse,
                        'history': final_history
                    }
                    
                    final_model_path = search_model_path  # 使用搜索階段的模型
                else:
                    # 最終訓練（Train+Val 重訓，使用遷移學習節省時間）
                    final_model, final_scores = final_training(
                        X_train, X_val, X_test, best_config, group_name, best_model=best_model
                    )
                    
                    # 保存最終模型
                    final_model_path = os.path.join(models_dir, f"{group_name}_final.h5")
                    final_model.save(final_model_path)
                
                # 繪製訓練歷史
                plot_path = os.path.join(plots_dir, f"{group_name}_training_history.png")
                plot_training_history(final_scores['history'], group_name, plot_path)
                
                # 保存標準化器
                scaler_path = os.path.join(models_dir, f"{group_name}_scaler.pkl")
                import pickle
                with open(scaler_path, 'wb') as f:
                    pickle.dump(scaler, f)
                
                # 壓縮並輸出時間序列資料（使用壓縮窗口的資料）
                print(f"\n[COMPRESS] 開始壓縮時間序列資料: {group_name} (窗口: {window_name})")
                compressed_data_path = compress_and_save_data(
                    final_model, scaler, compress_df_window, indicator_cols, group_name, 
                    window_output_dir, best_config['bottleneck'], window_name=window_name
                )
                
                # 儲存結果（按窗口和群組）
                if window_name not in all_results:
                    all_results[window_name] = {}
                
                all_results[window_name][group_name] = {
                    'input_dim': len(indicator_cols),
                    'best_config': best_config,
                    'search_results': [
                        {
                            'config': r['config'],
                            'train_mse': r['train_mse'],
                            'val_mse': r['val_mse'],
                            'history': r.get('history', {})  # 保存歷史記錄（包含時間信息）
                        }
                        for r in search_results
                    ],
                    'final_scores': final_scores,
                    'plot_path': plot_path,
                    'model_path': final_model_path,
                    'scaler_path': scaler_path,
                    'compressed_data_path': compressed_data_path
                }
                
                elapsed_group_time = (datetime.now() - window_start_time).total_seconds()
                print(f"[OK] {group_name} 處理完成！ (窗口內耗時: {elapsed_group_time:.1f}秒)")
                print(f"   已完成 {group_idx}/{total_groups} 個群組 ({group_idx/total_groups*100:.1f}%)")
                
            except Exception as e:
                print(f"[ERROR] {group_name} 處理失敗: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        # 為當前窗口保存結果到 Excel
        window_excel_path = save_results_to_excel(all_results[window_name], window_output_dir)
        print(f"[OK] 窗口 {window_name} 處理完成！")
        print(f"   結果已保存到: {window_output_dir}")
    
    # 保存整體 JSON 報告
    json_path = os.path.join(OUTPUT_DIR, f"all_windows_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
    # 移除無法序列化的對象
    json_results = {}
    for window_name, window_results in all_results.items():
        json_results[window_name] = {}
        for group_name, result in window_results.items():
            json_results[window_name][group_name] = {
                'input_dim': result['input_dim'],
                'best_config': result['best_config'],
                'final_scores': result['final_scores'],
                'search_results': result['search_results']
            }
    
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(json_results, f, indent=2, ensure_ascii=False)
    
    print(f"\n{'='*80}")
    print("[DONE] 所有滾動窗口處理完成！")
    print(f"[DIR] 結果目錄: {OUTPUT_DIR}")
    print(f"[JSON] 整體 JSON 報告: {json_path}")
    
    # 顯示所有窗口的壓縮資料路徑
    print(f"\n[COMPRESSED] 壓縮時間序列資料:")
    for window_name in all_results.keys():
        window_output_dir = os.path.join(OUTPUT_DIR, window_name)
        compressed_dir = os.path.join(window_output_dir, "compressed_data")
        if os.path.exists(compressed_dir):
            compressed_files = [f for f in os.listdir(compressed_dir) if f.endswith('.csv')]
            if compressed_files:
                print(f"   窗口 {window_name}:")
                print(f"     目錄: {compressed_dir}")
                print(f"     檔案數: {len(compressed_files)}")
                for f in compressed_files[:5]:  # 只顯示前5個
                    print(f"     - {f}")
                if len(compressed_files) > 5:
                    print(f"     ... 還有 {len(compressed_files) - 5} 個檔案")
    
    total_time = (datetime.now() - overall_start_time).total_seconds()
    print(f"\n[TIME] 總處理時間: {total_time:.1f}秒 ({total_time/60:.1f}分鐘)")
    print(f"[TIME] 結束時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 80)

if __name__ == "__main__":
    main()
